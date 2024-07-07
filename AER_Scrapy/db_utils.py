import sqlite3
import pandas as pd
import os
from utils import cprint
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class SQLiteDBHelper:
    def __init__(self, db_name, fresh_start=False, csv_path=None):
        self.db_name = db_name
        self.connection = self.create_connection()
        
        # To Do
        # if no tables found, print a msg then fresh_start=True

        if fresh_start:
            cprint('Deleting Tables')
            self.delete_tables() # Protected from None Deletion
            cprint('Creating Tables')
            self.create_tables() # Protected from Duplication
            cprint(f'Initializing Tables using {csv_path}')
            self.init_tables(csv_path)

    def create_tables(self):
        """Create the necessary tables for the schema."""
        create_polygon_table = """
        CREATE TABLE IF NOT EXISTS Polygon (
            id TEXT PRIMARY KEY
        );
        """
        create_landunit_table = """
        CREATE TABLE IF NOT EXISTS LandUnit (
            code TEXT PRIMARY KEY
        );
        """
        create_plotplan_table = """
        CREATE TABLE IF NOT EXISTS PlotPlan (
            name TEXT PRIMARY KEY,
            downloaded INTEGER DEFAULT 0
        );
        """
        create_polygon_landunit_table = """
        CREATE TABLE IF NOT EXISTS PolygonLandUnit (
            polygon_id TEXT,
            landunit_code TEXT,
            PRIMARY KEY (polygon_id, landunit_code)
            FOREIGN KEY (polygon_id) REFERENCES Polygon(id),
            FOREIGN KEY (landunit_code) REFERENCES LandUnit(code)
        );
        """
        create_landunit_plotplan_table = """
        CREATE TABLE IF NOT EXISTS LandUnitPlotPlan (
            landunit_code TEXT,
            plotplan_name TEXT,
            PRIMARY KEY (landunit_code, plotplan_name)
            FOREIGN KEY (landunit_code) REFERENCES LandUnit(code),
            FOREIGN KEY (plotplan_name) REFERENCES PlotPlan(name)
        );
        """
        
        self.execute_query(create_polygon_table)
        self.execute_query(create_landunit_table)
        self.execute_query(create_plotplan_table)
        self.execute_query(create_polygon_landunit_table)
        self.execute_query(create_landunit_plotplan_table)

    def init_tables(self, csv_path):
        """ Initialisation of tables Polygon, LandUnit, PolygonLandUnit."""
        self.import_poly_landunit_from_csv(csv_path)

    def import_poly_landunit_from_csv(self, csv_path):
        """Import PolyID and LandUnit values from a CSV file into the tables."""
        try:
            df = pd.read_csv(csv_path)
            for _, row in df.iterrows():
                polygon_id = row['PolyID']
                landunit_code = row['LandUnit']
                self.upsert_polygon(polygon_id)
                self.upsert_landunit(landunit_code)
                self.upsert_polygon_landunit(polygon_id, landunit_code)
        except Exception as e:
            cprint(f"Error importing CSV to database: {e}")
    
    def export_to_xlsx(self, excel_path):
        # List of table names
        tables = ['Polygon', 'LandUnit', 'PlotPlan', 'PolygonLandUnit', 'LandUnitPlotPlan']
        
        # Create a Pandas Excel writer using openpyxl
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
             # Get PlotPlan statistics
            plotplan_df = pd.read_sql_query("SELECT * FROM PlotPlan", self.connection)
            plotplans_detected = len(plotplan_df)
            plotplans_missing = len(plotplan_df[plotplan_df['downloaded'] == 0])
            plotplans_downloaded = len(plotplan_df[plotplan_df['downloaded'] == 1])
            plotplans_missing_percentage = (plotplans_missing / plotplans_detected) * 100
            plotplans_downloaded_percentage = (plotplans_downloaded / plotplans_detected) * 100

            # Create a DataFrame for the statistics
            stats_data = {
                'Metric': [
                    'PlotPlans Detected', 
                    'PlotPlans Downloaded', 
                    'PlotPlans Missing', 
                    'Percentage of PlotPlans Downloaded', 
                    'Percentage of PlotPlans Missing'
                ],
                'Value': [
                    plotplans_detected,
                    plotplans_downloaded,
                    plotplans_missing, 
                    f"{plotplans_downloaded_percentage:.2f}%",
                    f"{plotplans_missing_percentage:.2f}%", 
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            
            # Write the statistics DataFrame to a new sheet
            stats_df.to_excel(writer, sheet_name='PlotPlanStatistics', index=False)

            # Perform the join query
            join_query = """
            SELECT pl.polygon_id, pl.landunit_code, lu.plotplan_name
            FROM PolygonLandUnit pl
            JOIN LandUnitPlotPlan lu ON pl.landunit_code = lu.landunit_code;
            """
            join_df = pd.read_sql_query(join_query, self.connection)
            # Write the joined DataFrame to a new sheet
            join_df.to_excel(writer, sheet_name='JoinedData', index=False)
            
            for table in tables:
                # Read each table into a DataFrame
                df = pd.read_sql_query(f"SELECT * FROM {table}", self.connection)
                # Write the DataFrame to a sheet named after the table
                df.to_excel(writer, sheet_name=table, index=False)

        # Load the workbook to adjust column widths
        workbook = load_workbook(excel_path)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the adjusted workbook
        workbook.save(excel_path)
    
    def export_to_csv(self):
        pass
    
    def export_to_txt(self):
        pass

    def query_polygon(self, landunit_code):
        """Query and return the polygon_id associated with a given landunit_code."""
        query = """
        SELECT polygon_id 
        FROM PolygonLandUnit
        WHERE landunit_code = ?
        """
        result = self.execute_query(query, (landunit_code,))
        if result:
            return result[0][0]
        else:
            cprint(f"No polygon found for landunit: {landunit_code}")
            return None
        
    def upsert_polygon(self, polygon_id):
        query = "INSERT OR REPLACE INTO Polygon (id) VALUES (?)"
        self.execute_query(query, (polygon_id,))

    def upsert_landunit(self, landunit_code):
        query = "INSERT OR REPLACE INTO LandUnit (code) VALUES (?)"
        self.execute_query(query, (landunit_code,))

    def upsert_plotplan(self, plotplan_name, downloaded=False):
        if downloaded:
            query = "INSERT OR REPLACE INTO PlotPlan (name, downloaded) VALUES (?, ?)"
            self.execute_query(query, (plotplan_name,1))
        else:
            query = "INSERT OR REPLACE INTO PlotPlan (name) VALUES (?)"
            self.execute_query(query, (plotplan_name,))

    def upsert_polygon_landunit(self, polygon_id, landunit_code):
        query = "INSERT OR REPLACE INTO PolygonLandUnit (polygon_id, landunit_code) VALUES (?, ?)"
        self.execute_query(query, (polygon_id, landunit_code))

    def upsert_landunit_plotplan(self, landunit_code, plotplan_name):
        query = "INSERT OR REPLACE INTO LandUnitPlotPlan (landunit_code, plotplan_name) VALUES (?, ?)"
        self.execute_query(query, (landunit_code, plotplan_name))

    def delete_tables(self):
        """Drop all tables in the database."""
        drop_polygon_table = "DROP TABLE IF EXISTS Polygon"
        drop_landunit_table = "DROP TABLE IF EXISTS LandUnit"
        drop_plotplan_table = "DROP TABLE IF EXISTS PlotPlan"
        drop_polygon_landunit_table = "DROP TABLE IF EXISTS PolygonLandUnit"
        drop_landunit_plotplan_table = "DROP TABLE IF EXISTS LandUnitPlotPlan"
        self.execute_query(drop_polygon_table)
        self.execute_query(drop_landunit_table)
        self.execute_query(drop_plotplan_table)
        self.execute_query(drop_polygon_landunit_table)
        self.execute_query(drop_landunit_plotplan_table)

    def execute_query(self, query, params=None):
        """Wrapper to execute a query and return the results."""
        try:
            cursor = self.connection.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            self.connection.commit()
            return cursor.fetchall()
        except sqlite3.Error as e:
            cprint(f"Error executing query: {e}")
            return None

    def create_connection(self):
        """Create a database connection and return the connection object."""
        try:
            conn = sqlite3.connect(self.db_name)
            return conn
        except sqlite3.Error as e:
            cprint(f"Error creating connection to database: {e}")
            return None
    
    def close_connection(self):
        """Close the database connection."""
        if self.connection:
            self.connection.close()

    def debug(self):
        """Print out all tables and their contents."""
        tables = self.execute_query("SELECT name FROM sqlite_master WHERE type='table';")
        for table in tables:
            table_name = table[0]
            print(f"\nTable: {table_name}")
            rows = self.execute_query(f"SELECT * FROM {table_name}")
            for row in rows:
                print(row)